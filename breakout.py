from collections import namedtuple
from gettext import translation
from nntplib import NNTP
import os
import sys

import torch
from torch import nn
import torch.nn.functional as F
import matplotlib.pyplot as plt
import numpy as np
import collections
import gym
from gym import spaces
from collections import deque
import random
import torch.optim as optim
import math
import cv2
#from skimage import transform
from gym.spaces import Box
from gym.wrappers import FrameStack
from torchvision import transforms as T
from PIL import Image
import openpyxl

os.environ["KMP_DUPLICATE_LIB_OK"]="TRUE"
# Check if GPU is avalible
torch.cuda.empty_cache()
device = torch.device("cuda" if torch.cuda.is_available() else "cpu")

print(f'Using {device} device')

# Defining classes and functions ---------------------------------------------------------------

class NeuralNetwork(nn.Module):
    # Creating Neural network with two hidden layers
    # Input: 4
    # Output 2
    def __init__(self,state_size,action_size):
        super(NeuralNetwork, self).__init__()
        self.conv1 = nn.Conv2d(4, 32, 8, 4)
        self.conv2 = nn.Conv2d(32, 64, 4, 2)
        self.conv3 = nn.Conv2d(64, 64, 3, 1)
        self.fc1 = nn.Linear(7 * 7 * 64, 512)
        self.fc2 = nn.Linear(512, action_size)
        
    # Defining forward function
    def forward(self,x):
        x = F.relu(self.conv1(x))
        x = F.relu(self.conv2(x))
        x = F.relu(self.conv3(x))
        x = x.view(-1, 7 * 7 * 64)
        x = F.relu(self.fc1(x))
        x = self.fc2(x)
        return x


# datatype for storing multiple things
Experience = namedtuple('Experience',
                        ('state','action','reward','next_state'))

#Create a memory class
class Memory:
    def __init__(self, capacity):
        self.memory = deque([],maxlen=capacity)     #deque is like a list but fast

    def __len__(self):            
        return len(self.memory)

    def cache(self, state, next_state, action, reward, done):
        """
        Store the experience to self.memory (replay buffer)

        Inputs:
        state (LazyFrame),
        next_state (LazyFrame),
        action (int),
        reward (float),
        done(bool))
        """
        state = state.__array__()
        next_state = next_state.__array__()

        state = torch.tensor(state)
        next_state = torch.tensor(next_state)
        action = torch.tensor([action])
        reward = torch.tensor([reward])
        done = torch.tensor([done])
        
        self.memory.append((state, next_state, action, reward, done,))
    
    def sample(self,batch_size:int):
        return random.sample(self.memory,batch_size)
    
    def recall(self):
        """
        Retrieve a batch of experiences from memory
        """
        batch = random.sample(self.memory, batch_size)
        state, next_state, action, reward, done = map(torch.stack, zip(*batch))
        return state, next_state, action.squeeze(), reward.squeeze(), done.squeeze()


# Wrappers from https://github.com/jorditorresBCN/Deep-Reinforcement-Learning-Explained/blob/master/DRL_15_16_17_DQN_Pong.ipynb

class FireResetEnv(gym.Wrapper):
    def __init__(self, env=None):
        super(FireResetEnv, self).__init__(env)
        assert env.unwrapped.get_action_meanings()[1] == 'FIRE'
        print(env.unwrapped.get_action_meanings())
        assert len(env.unwrapped.get_action_meanings()) >= 3

    def step(self, action):
        return self.env.step(action)

    def reset(self):
        self.env.reset()
        obs, _, done, _ = self.env.step(1)
        if done:
            self.env.reset()
        # obs, _, done, _ = self.env.step(2) #Should not be included
        # if done:
        #     self.env.reset()
        return obs

class MaxAndSkipEnv(gym.Wrapper):
    def __init__(self, env=None, skip=4):
        super(MaxAndSkipEnv, self).__init__(env)
        # most recent raw observations (for max pooling across time steps)
        self._obs_buffer = collections.deque(maxlen=2)
        self._skip = skip

    def step(self, action):
        total_reward = 0.0
        done = None
        for _ in range(self._skip):
            obs, reward, done, info = self.env.step(action)
            self._obs_buffer.append(obs)
            total_reward += reward
            if done:
                break
        max_frame = np.max(np.stack(self._obs_buffer), axis=0)
        return max_frame, total_reward, done, info

    def reset(self):
        self._obs_buffer.clear()
        obs = self.env.reset()
        self._obs_buffer.append(obs)
        return obs


class ProcessFrame84(gym.ObservationWrapper):
    def __init__(self, env=None):
        super(ProcessFrame84, self).__init__(env)
        self.observation_space = gym.spaces.Box(low=0, high=255, shape=(84, 84, 1), dtype=np.uint8)

    def observation(self, obs):
        return ProcessFrame84.process(obs)

    @staticmethod
    def process(frame):
        
    
        if frame.size == 210 * 160 * 3:
            img = np.reshape(frame, [210, 160, 3]).astype(np.float32)
        elif frame.size == 250 * 160 * 3:
            img = np.reshape(frame, [250, 160, 3]).astype(np.float32)
        else:
            assert False, "Unknown resolution."
        


        img1 = img[:, :, 0] * 1 + img[:, :, 1] * 1 + img[:, :, 2] * 1 #Tar bort färger
        #cv2.imwrite("img1.jpg",img1)
        img2 = img1[32:192, 8:152]  # Cropping from 210x160 to 160x144
        #cv2.imwrite("img2.jpg",img2)
        img3 = cv2.resize(img2, (84, 84), interpolation=cv2.INTER_AREA) # Compressing from 160x144 to 84x84
        #cv2.imwrite("img3.jpg",img3)

        th, img4 = cv2.threshold(src=img3, thresh=177, maxval=255, type=cv2.THRESH_BINARY) # Binary black and white
        #cv2.imwrite("img4.jpg",img4)
        img5 = np.reshape(img4, [84, 84, 1])
        #cv2.imwrite("img5.jpg",img5)

        #img6 = img5.astype(np.uint8)
        #cv2.imwrite("img6.jpg",img6)

        return img5.astype(np.uint8)
class BufferWrapper(gym.ObservationWrapper):
    def __init__(self, env, n_steps, dtype=np.float32):
        super(BufferWrapper, self).__init__(env)
        self.dtype = dtype
        old_space = env.observation_space
        self.observation_space = gym.spaces.Box(old_space.low.repeat(n_steps, axis=0),
                                                old_space.high.repeat(n_steps, axis=0), dtype=dtype)

    def reset(self):
        self.buffer = np.zeros_like(self.observation_space.low, dtype=self.dtype)
        return self.observation(self.env.reset())

    def observation(self, observation):
        self.buffer[:-1] = self.buffer[1:]
        self.buffer[-1] = observation
        return self.buffer


class ImageToPyTorch(gym.ObservationWrapper):
    def __init__(self, env):
        super(ImageToPyTorch, self).__init__(env)
        old_shape = self.observation_space.shape
        self.observation_space = gym.spaces.Box(low=0.0, high=1.0, shape=(old_shape[-1], 
                                old_shape[0], old_shape[1]), dtype=np.float32)

    def observation(self, observation):
        return np.moveaxis(observation, 2, 0)


class ScaledFloatFrame(gym.ObservationWrapper):
    def observation(self, obs):
        return np.array(obs).astype(np.float32) / 255.0

class Discretizer(gym.ActionWrapper):
    """
    Wrap a gym environment and make it use discrete actions.
    Args:
        combos: ordered list of lists of valid button combinations
    """

    def __init__(self, env, combos):
        super().__init__(env)
        print(env.action_space)
        
        assert isinstance(env.action_space, gym.spaces.MultiBinary)
        buttons = env.unwrapped.buttons
        self._decode_discrete_action = []
        for combo in combos:
            arr = np.array([False] * env.action_space.n)
            for button in combo:
                arr[buttons.index(button)] = True
            self._decode_discrete_action.append(arr)

        self.action_space = gym.spaces.Discrete(len(self._decode_discrete_action))

    def action(self, act):
        return self._decode_discrete_action[act].copy()


class EpisodicLifeEnv(gym.Wrapper):
    def __init__(self, env):
        """Make end-of-life == end-of-episode, but only reset on true game over.
        Done by DeepMind for the DQN and co. since it helps value estimation.
        """
        gym.Wrapper.__init__(self, env)
        self.lives = 0
        self.was_real_done  = True

    def step(self, action):
        obs, reward, done, info = self.env.step(action)
        self.was_real_done = done
        # check current lives, make loss of life terminal,
        # then update lives to handle bonus lives
        lives = self.env.unwrapped.ale.lives()
        if lives < self.lives and lives > 0:
            # for Qbert sometimes we stay in lives == 0 condition for a few frames
            # so it's important to keep lives > 0, so that we only reset once
            # the environment advertises done.
            done = True
        self.lives = lives
        return obs, reward, done, info

    def reset(self, **kwargs):
        """Reset only when lives are exhausted.
        This way all states are still reachable even though lives are episodic,
        and the learner need not know about any of this behind-the-scenes.
        """
        if self.was_real_done:
            obs = self.env.reset(**kwargs)
        else:
            # no-op step to advance from terminal/lost life state
            obs, _, _, _ = self.env.step(0)
        self.lives = self.env.unwrapped.ale.lives()
        return obs

def make_env(env_name):
    env = gym.make(env_name)
    env = MaxAndSkipEnv(env)
    env = EpisodicLifeEnv(env)
    env = FireResetEnv(env)
    env = ProcessFrame84(env)
    env = ImageToPyTorch(env)
    env = BufferWrapper(env, 4)
    return ScaledFloatFrame(env)

env_name = "BreakoutNoFrameskip-v4"

env = make_env(env_name)




#Take action function 
def take_action(state):
    # global steps_done
    # eps_threshold = EPS_END + (EPS_START - EPS_END) * \
    #     math.exp(-1. * steps_done / EPS_DECAY)
    # steps_done += 1

    eps_threshold = (EPS_END +
                max(0, (EPS_START - EPS_END) * (EPS_DECAY -
                max(0, numSteps - learn_start))/EPS_DECAY))

    if numSteps % 1000 == 0:
        print('EPS: ',eps_threshold)

    #takes a random action if under eps_threshold, uses the model otherwise
    if random.uniform(0,1) < eps_threshold:
        action = random.randrange(action_size)
    else:
        state = state.__array__()
        state = torch.tensor(state).to(device)
        state = state.unsqueeze(0)
        action_values = model(state)
        action = torch.argmax(action_values, axis=1).item()
        del state
    return action


def optimize_model():
    if len(memory) < learn_start:
        return

    # Samples random batch from memory    
    state, next_state, action, reward, done = memory.recall()

    if torch.cuda.is_available():
        state = state.cuda()
        next_state = next_state.cuda()
        action = action.cuda()
        reward = reward.cuda()
        done = done.cuda()

    state_action_values = model(state)[np.arange(0, batch_size), action]       
    
    next_state_Q = model(next_state)
    best_action = torch.argmax(next_state_Q, axis=1)
    next_Q = target_model(next_state)[np.arange(0, batch_size), best_action]
    expected_state_action_values = (reward + (1 - done.float()) * gamma * next_Q).float()
    
    # Setup the loss function using the mean square error
    #criterion = nn.MSELoss()
    criterion = nn.HuberLoss()
    #criterion = nn.SmoothL1Loss()
    
    
    loss = criterion(state_action_values,expected_state_action_values.detach())
    
    #Reset gradient, calculate new gradient and take a step
    optimizer.zero_grad()
    loss.backward()
    # for param in model.parameters():
    #     param.grad.data.clamp_(-1, 1)
    optimizer.step()

    del state, next_state, action, reward, done
    
def plot_data(average_scores_epoch, max_q_mean_epoch, epoch):
    plt.figure(1)
    plt.plot(average_scores_epoch,'b--')
    plt.xlabel('Epoch')
    plt.ylabel("score")
    title1 = "Final average score: " + str(np.mean(average_scores_epoch[-100:])) + ", For epoch: " + str(epoch)
    plt.title(title1)
    filename1 = "average_score.png"
    plt.savefig(filename1)
    plt.clf()

    plt.figure(2)
    plt.plot(max_q_mean_epoch,'b--')
    plt.xlabel('Epoch')
    plt.ylabel("average max value of Q")
    title2 = "Max_q_mean for epoch: " + str(epoch)
    plt.title(title2)
    filename2 = "max_q_mean.png"
    plt.savefig(filename2)
    plt.clf()

def create_excel(scores, max_q_mean, average_scores, i_episode):
    exceldokument = openpyxl.Workbook()

    # ws_scores = 'Scores'
    # ws_average = 'Average Scores'
    # ws_q_value = 'Max Q value'

    ws_main = 'data'

    # exceldokument.create_sheet(ws_scores)
    # exceldokument.create_sheet(ws_average)
    # exceldokument.create_sheet(ws_q_value)
    exceldokument.create_sheet(ws_main)
    
    # ws1 = exceldokument.get_sheet_by_name(ws_scores)
    # ws2 = exceldokument.get_sheet_by_name(ws_average)
    # ws3 = exceldokument.get_sheet_by_name(ws_q_value)
    # max_q_mean_c = [item for sublist in max_q_mean for item in sublist]

    exceldokument[ws_main].cell(row=1,column=1).value = 'scores'
    exceldokument[ws_main].cell(row=1,column=2).value = 'average_scores'
    exceldokument[ws_main].cell(row=1,column=3).value = 'max_q_mean'

    for index in range(i_episode):
        
        # ws1.cell(row=index+1,column=i+1).value = scores[index]
        # ws2.cell(row=index+1,column=i+1).value = average_scores[index]
        # ws3.cell(row=index+1,column=i+1).value = max_q_mean_c[index]
        
        exceldokument[ws_main].cell(row=index+2,column=1).value = scores[index]
        exceldokument[ws_main].cell(row=index+2,column=2).value = average_scores[index]
        exceldokument[ws_main].cell(row=index+2,column=3).value = max_q_mean[index]
        #print(scores[index])
    filename = 'data_' + str(i_episode) + '.xlsx'
    exceldokument.save(filename)

#Initializing starts here ----------------------------------

# #env = gym.make('BreakoutNoFrameskip-v4', render_mode='human')
# env = gym.make('BreakoutNoFrameskip-v4')
# #env = gym.make('ALE/Breakout-v5')
# #env = ale.loadROM('Breakout')
# env.reset()
# # Apply Wrappers to environment
# env = SkipFrame(env, skip=4)
# env = GrayScaleObservation(env)
# env = ResizeObservation(env, shape=84)
# env = FrameStack(env, num_stack=4)

#Constants  
batch_size = 32
TARGET_UPDATE = 10000
UPDATE_FREQ = 4
gamma = 0.99
EPISODES = 100000                          
test_states_no = 1000
steps_done = 0
EPS_START = 1
EPS_END = 0.05
EPS_DECAY = 1000000
learn_start = 5000
numSteps = 0
num_training_step = 0
learning_rate = 0.00001
epoch = 0
END_EPOCH = 100
frames_per_EPOCH = 50000
i_episode = 0
last_i_episode = 0
trained = False

state_size = env.observation_space.shape[0]
action_size = env.action_space.n

#Create replay memory
memory = Memory(100000)


#Creates model and target model
model = NeuralNetwork(state_size,action_size).to(device)
target_model = NeuralNetwork(state_size,action_size).to(device)
target_model.load_state_dict(model.state_dict())
target_model.eval()


#Using the Adam optimizer
optimizer = optim.Adam(model.parameters(), lr=learning_rate)




if __name__ == "__main__":

    #create test states for plotting q-values
    test_states = torch.empty(test_states_no,4,84,84, requires_grad=False).to(device)
    
    #max_q = np.zeros((EPISODES, test_states_no))
    #max_q_mean = np.zeros((EPISODES,1))
    
    done = True
    for i in range(test_states_no):
        if done:
            done = False
            state = env.reset()
            state = state.__array__()
            state = torch.tensor(state, requires_grad=False).to(device)
            test_states[i] = state
        else:
            action = random.randrange(action_size)
            next_state, reward, done, info = env.step(action)
            next_state = next_state.__array__()
            next_state = torch.tensor(next_state, requires_grad=False).to(device)
            state = next_state
    
    scores, episodes, average_scores, max_q_mean = [], [], [], []
    average_scores_epoch, max_q_mean_epoch, num_episode = [], [], []
    #The main for loop where we go through the whole range of Episodes
    while epoch < END_EPOCH:

        #calculate mean q-value
        done = False
        score = 0
        
        tmp = model(test_states)
        # max_q[i_episode][:] = np.max(tmp.cpu().detach().numpy(), axis=1)
        # max_q_mean[i_episode] = np.mean(max_q[i_episode][:])
        max_q_mean.append(np.mean(np.max(tmp.cpu().detach().numpy(), axis=1)))
        
        #Reset env
        next_state = env.reset()

        while not done:
            #env.render() #Can remove when training
            
            state = next_state
            
            action = take_action(state) 
            numSteps += 1
            #get feedback from env
            next_state, reward, done, info = env.step(action)

            score += reward
            memory.cache(state, next_state, action, reward, done)
                        
            #Train the model
            if numSteps % UPDATE_FREQ == 0:
                num_training_step += 1
                trained = True
                optimize_model()
            
            #Updates the target network
            if numSteps % TARGET_UPDATE == 1:
                print('-----------model update------------')
                target_model.load_state_dict(model.state_dict())

            if num_training_step % frames_per_EPOCH == 0 and num_training_step>0 and trained:
                length_epoch = i_episode - last_i_episode
                last_i_episode = i_episode
                
                average_scores_epoch.append(np.mean(average_scores[-length_epoch:]))
                max_q_mean_epoch.append(np.mean(max_q_mean[-length_epoch:]))
                torch.save(model.state_dict(),'Models\model_epoch_'+str(epoch)+'.pth')
                trained = False
                plot_data(average_scores_epoch,max_q_mean_epoch,epoch)
                epoch += 1

            #If the episode is finished
            if done:
                scores.append(score)
                average_scores.append(np.mean(scores[-100:]))
                print("episode:", i_episode, "  score:", score," q_value:", max_q_mean[i_episode],
                "  memory length:", len(memory), " average score: ",round(np.mean(scores[-100:]),2))
                i_episode += 1
                break


plot_data(average_scores_epoch,max_q_mean_epoch,epoch)
create_excel(scores, max_q_mean, average_scores,epoch)

torch.save(model.state_dict(),'Models\model_final.pth')   

